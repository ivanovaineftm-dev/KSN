from __future__ import annotations

from datetime import date, datetime
from difflib import get_close_matches
from pathlib import Path
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill


TARGET_ROLE = "стажер"
INVALID_ROW_FILL = PatternFill(fill_type="solid", start_color="FFFFC7CE", end_color="FFFFC7CE")
NORMALIZED_DEPARTMENT_COLUMN = "Подразделение (нормализованное)"
NOT_FOUND_LABEL = "Не найдено"

MENTOR_ROLE_RULES: dict[str, set[str]] = {
    "бариста-стажер": {"бариста"},
    "кассир-стажер": {"кассир", "старший кассир", "повар-универсал"},
    "старший кассир-стажер": {"старший кассир", "заместитель директора"},
    "повар-стажер": {
        "повар",
        "повар-универсал",
        "повар-бригадир",
        "повар 1 категории",
        "повар 2категории",
    },
    "повар-универсал стажер": {"повар-универсал", "повар", "старший кассир", "кассир"},
    "работник торгового зала-стажер": {
        "кассир",
        "старший кассир",
        "работник торгового зала",
        "повар-универсал",
    },
    "оператор-кассир стажер": {"оператор-кассир", "старший оператор-кассир"},
    "заместитель директора азс по направлению кафе-стажер": {
        "заместитель директора азс по направлению кафе"
    },
    "оператор мини-кафе-стажер": {
        "оператор мини-кафе",
        "старший оператор-кассир",
        "оператор-кассир",
    },
    "бармен-бариста-стажер": {"бармен-бариста", "менеджер-кассир"},
    "менеджер-кассир стажер": {"менеджер-кассир", "заместитель директора"},
    "официант-стажер": {"официант", "менеджер кассир"},
    "хостес-стажер": {"хостес", "менеджер кассир"},
}


def _is_blank(value: object) -> bool:
    if pd.isna(value):
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _contains_february(value: object) -> bool:
    if _is_blank(value):
        return False

    if isinstance(value, (datetime, date)):
        return value.month == 2

    value_text = str(value).strip().lower()
    if "феврал" in value_text:
        return True

    return re.search(r"(?:^|\D)\d{1,2}\.02\.\d{4}(?:$|\D)", value_text) is not None


def _contains_target_role(value: object) -> bool:
    if _is_blank(value):
        return False

    value_text = str(value).strip().lower().replace("ё", "е")
    normalized_text = value_text.replace("–", "-").replace("—", "-")

    return TARGET_ROLE in normalized_text


def _normalize_role(value: object) -> str:
    if _is_blank(value):
        return ""

    return (
        str(value)
        .strip()
        .lower()
        .replace("ё", "е")
        .replace("–", "-")
        .replace("—", "-")
        .replace(" - ", "-")
        .replace("- ", "-")
        .replace(" -", "-")
        .replace("  ", " ")
    )


def _normalize_text(value: object) -> str:
    if _is_blank(value):
        return ""

    return re.sub(r"\s+", " ", str(value)).strip().lower()


def _normalize_department_key(value: object) -> str:
    return _normalize_text(value)


def _normalize_department_display(value: object) -> str:
    return re.sub(r"\s+", " ", str(value)).strip()


def _mentor_role_is_valid(trainee_role: object, mentor_role: object) -> bool:
    normalized_trainee_role = _normalize_role(trainee_role)
    allowed_mentor_roles = MENTOR_ROLE_RULES.get(normalized_trainee_role)
    if not allowed_mentor_roles:
        return True

    normalized_mentor_role = _normalize_role(mentor_role)
    if not normalized_mentor_role:
        return False

    return normalized_mentor_role in allowed_mentor_roles


def _row_has_mentor_validation_error(trainee_role: object, mentor_role: object) -> bool:
    return _is_blank(mentor_role) or not _mentor_role_is_valid(trainee_role, mentor_role)


def _read_excel_file(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return pd.read_excel(path, engine="openpyxl")
    if suffix == ".xls":
        return pd.read_excel(path, engine="xlrd")
    raise ValueError(f"Неподдерживаемый формат файла: {path.suffix}")


def _build_department_dictionary(locations_path: Path | None) -> dict[str, str]:
    if locations_path is None:
        return {}

    locations_df = _read_excel_file(locations_path)
    if locations_df.shape[1] < 2:
        return {}

    result: dict[str, str] = {}
    for value in locations_df.iloc[:, 1]:
        if _is_blank(value):
            continue
        display = _normalize_department_display(value)
        result[_normalize_department_key(value)] = display
    return result


def _build_barista_dictionary(barista_path: Path | None) -> dict[str, str]:
    if barista_path is None:
        return {}

    barista_df = _read_excel_file(barista_path)
    if barista_df.shape[1] < 3:
        return {}

    result: dict[str, str] = {}
    for row in barista_df.iloc[:, [1, 2]].itertuples(index=False, name=None):
        department_value, search_value = row
        if _is_blank(department_value) or _is_blank(search_value):
            continue
        result[_normalize_text(search_value)] = _normalize_department_display(department_value)
    return result


def _match_department(department_value: object, locations: dict[str, str]) -> str:
    key = _normalize_department_key(department_value)
    if not key:
        return NOT_FOUND_LABEL

    exact = locations.get(key)
    if exact:
        return exact

    location_keys = list(locations.keys())
    for location_key in location_keys:
        if key in location_key or location_key in key:
            return locations[location_key]

    fuzzy = get_close_matches(key, location_keys, n=1, cutoff=0.78)
    if fuzzy:
        return locations[fuzzy[0]]

    return NOT_FOUND_LABEL


def _is_barista_role(value: object) -> bool:
    return _normalize_role(value) == "бариста"


def _match_barista_department(mentor_value: object, barista_dictionary: dict[str, str]) -> str | None:
    mentor_key = _normalize_text(mentor_value)
    if not mentor_key:
        return None

    exact_match = barista_dictionary.get(mentor_key)
    if exact_match:
        return exact_match

    dictionary_keys = list(barista_dictionary.keys())
    for item_key in dictionary_keys:
        if mentor_key in item_key or item_key in mentor_key:
            return barista_dictionary[item_key]

    fuzzy = get_close_matches(mentor_key, dictionary_keys, n=1, cutoff=0.78)
    if fuzzy:
        return barista_dictionary[fuzzy[0]]

    return None


def process_excel(
    input_path: Path,
    locations_path: Path | None,
    barista_path: Path | None,
    output_path: Path,
) -> list[dict[str, int | str]]:
    main_df = _read_excel_file(input_path)
    department_dictionary = _build_department_dictionary(locations_path)
    barista_dictionary = _build_barista_dictionary(barista_path)

    if main_df.shape[1] < 8:
        raise ValueError("В основном файле недостаточно столбцов для обработки (ожидается минимум 8).")

    processed_df = main_df.copy()
    replacement_mask = processed_df.iloc[:, 5].apply(_is_barista_role)
    replacement_values = processed_df.loc[replacement_mask, :].apply(
        lambda row: _match_barista_department(row.iloc[4], barista_dictionary),
        axis=1,
    )
    for row_index, replacement_department in replacement_values.items():
        if replacement_department:
            processed_df.iat[row_index, 3] = replacement_department

    processed_df[NORMALIZED_DEPARTMENT_COLUMN] = processed_df.iloc[:, 3].apply(
        lambda value: _match_department(value, department_dictionary)
    )

    keep_mask = processed_df.iloc[:, 2].apply(_contains_target_role)
    keep_mask &= ~processed_df.iloc[:, 7].apply(_is_blank)
    keep_mask &= ~processed_df.iloc[:, 7].apply(_contains_february)
    processed_df = processed_df[keep_mask].copy()

    invalid_mask = processed_df.apply(
        lambda row: _row_has_mentor_validation_error(row.iloc[2], row.iloc[5]),
        axis=1,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Processed"
    sheet.append(processed_df.columns.tolist())

    for row_offset, row_values in enumerate(processed_df.itertuples(index=False, name=None), start=2):
        sheet.append(list(row_values))
        if bool(invalid_mask.iloc[row_offset - 2]):
            for col_idx in range(1, len(processed_df.columns) + 1):
                sheet.cell(row=row_offset, column=col_idx).fill = INVALID_ROW_FILL

    workbook.save(output_path)

    department_stats: dict[str, dict[str, int | str]] = {}
    for row_values, has_error in zip(processed_df.itertuples(index=False, name=None), invalid_mask.tolist()):
        department_name = row_values[-1]
        if department_name == NOT_FOUND_LABEL:
            continue

        department_key = _normalize_department_key(department_name)
        stats = department_stats.setdefault(
            department_key,
            {"department": str(department_name), "total_rows": 0, "valid_rows": 0},
        )
        stats["total_rows"] += 1
        if not has_error:
            stats["valid_rows"] += 1

    analytics: list[dict[str, int | str]] = []
    for stats in department_stats.values():
        total_rows = int(stats["total_rows"])
        valid_rows = int(stats["valid_rows"])
        quality = round((valid_rows / total_rows) * 100) if total_rows else 0
        analytics.append(
            {
                "department": str(stats["department"]),
                "total_rows": total_rows,
                "valid_rows": valid_rows,
                "quality": quality,
            }
        )

    analytics.sort(key=lambda item: (-int(item["quality"]), str(item["department"])))
    return analytics
